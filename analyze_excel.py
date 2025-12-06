import openpyxl
import json

# Load the workbook
wb = openpyxl.load_workbook('AET10-SOFTWARE.xlsx', data_only=True)

print(f"Available sheets: {wb.sheetnames}\n")

output_data = {}

# Read each sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}\n")
    
    sheet_data = []
    
    # Print all rows
    for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        # Filter out completely empty rows
        if any(cell is not None for cell in row):
            row_data = [str(cell) if cell is not None else '' for cell in row]
            sheet_data.append(row_data)
            print(f"Row {idx}: {row_data}")
    
    output_data[sheet_name] = sheet_data
    print(f"\nTotal rows in {sheet_name}: {len(sheet_data)}\n")

# Save to JSON for easier parsing
with open('excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(output_data, f, indent=2, ensure_ascii=False)

print("\nData saved to excel_data.json")
