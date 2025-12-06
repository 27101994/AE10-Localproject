import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('AET10-SOFTWARE.xlsx', data_only=True)

print(f"Available sheets: {wb.sheetnames}\n")

# Read each sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f"\n{'='*80}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*80}\n")
    
    # Print all rows (limit to first 100 rows to avoid too much output)
    for idx, row in enumerate(sheet.iter_rows(max_row=100, values_only=True), 1):
        # Filter out completely empty rows
        if any(cell is not None for cell in row):
            print(f"Row {idx}: {row}")
    
    print("\n")
